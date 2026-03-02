"""
app.py — PropTech AVM: Vision Extractor (Production)
-----------------------------------------------------
A Streamlit dashboard that ingests scanned Indian property documents (PDFs
with printed + handwritten Hindi), renders ALL pages as high-resolution
images via PyMuPDF, and sends them to Google Gemini Vision to extract
structured property data.

API key is loaded from .env file — end users never touch it.
"""

import io
import json
import os
import re
import traceback
import zipfile


import fitz  # PyMuPDF
from dotenv import load_dotenv
from google import genai
import streamlit as st
from PIL import Image
import openpyxl
import requests
import folium
from streamlit_folium import st_folium

# ── Load configuration ────────────────────────────────────────────────────────
# Priority: st.secrets (Streamlit Cloud) > .env file (local) > empty
load_dotenv()


def _get_secret(key: str, default: str = "") -> str:
    """Read from Streamlit secrets first, then environment variables."""
    try:
        return st.secrets[key]
    except (KeyError, FileNotFoundError):
        return os.getenv(key, default)


GEMINI_API_KEY = _get_secret("GEMINI_API_KEY")
DEFAULT_MODEL = _get_secret("GEMINI_MODEL", "gemini-2.5-flash")
GMAPS_API_KEY = _get_secret("GMAPS_API_KEY")
MAPBOX_API_KEY = _get_secret("MAPBOX_API_KEY")

# ─────────────────────────────────────────────────────────────────────────────
# Page configuration
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="docfinder: Vision Extractor",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# Custom CSS — dark, premium PropTech aesthetic
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <style>
    /* ── DARK MODE METRIC CARD DASHBOARD ── */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }

    /* Remove default white backgrounds from containers */
    [data-testid="stAppViewContainer"] { background-color: #0E1117; }
    [data-testid="stHeader"] { background-color: rgba(0,0,0,0); }
    [data-testid="stToolbar"] { right: 2rem; }
    
    /* Remove ugly borders from expanders and tables */
    div[data-testid="stExpander"] { 
        border: none; 
        box-shadow: none; 
        background-color: #161B22; 
        border-radius: 10px; 
    }

    /* Reduce Streamlit's huge default horizontal spacing on desktop, and tighten on mobile */
    .block-container {
        max-width: 1200px !important;
        padding-top: 2rem !important;
        padding-bottom: 2rem !important;
        padding-left: clamp(1rem, 5vw, 3rem) !important;
        padding-right: clamp(1rem, 5vw, 3rem) !important;
    }

    /* Status badge - Dark Mode */
    .badge {
        display: inline-block;
        padding: 0.25rem 0.85rem;
        border-radius: 20px;
        font-size: 0.72rem;
        font-weight: 700;
        letter-spacing: 0.05em;
        text-transform: uppercase;
        margin-bottom: 0.6rem;
    }
    .badge-blue  { background: #1e3a8a; color: #bfdbfe; border: 1px solid #3b82f6; }
    .badge-green { background: #064e3b; color: #a7f3d0; border: 1px solid #10b981; }
    .badge-amber { background: #78350f; color: #fde68a; border: 1px solid #f59e0b; }
    .badge-red   { background: #7f1d1d; color: #fecaca; border: 1px solid #ef4444; }

    /* Buttons */
    .stButton > button {
        background: #3182ce;
        color: white; border: none; border-radius: 10px;
        padding: 0.55rem 1.5rem; font-weight: 600;
        transition: opacity 0.18s, transform 0.12s, box-shadow 0.2s;
        box-shadow: 0 2px 4px rgba(49, 130, 206, 0.3);
    }
    .stButton > button:hover { 
        opacity: 0.9; transform: translateY(-1px);
        box-shadow: 0 4px 6px rgba(49, 130, 206, 0.4);
        color: white;
    }

    /* Headings - Dark Mode & responsive */
    h1 { color: #f8fafc !important; font-weight: 700 !important; font-size: clamp(2rem, 5vw, 2.5rem) !important; }
    h2 { color: #e2e8f0 !important; font-weight: 600 !important; font-size: clamp(1.5rem, 4vw, 2rem) !important; }
    h3 { color: #cbd5e0 !important; font-size: clamp(1.2rem, 3vw, 1.5rem) !important;}
    p, span, div { color: #cbd5e0; }



    /* Alerts */
    .stSuccess { background: #f0fff4 !important; border: 1px solid #c6f6d5 !important; border-radius: 8px; color: #276749 !important; }
    .stWarning { background: #fffaf0 !important; border: 1px solid #feebc8 !important; border-radius: 8px; color: #c05621 !important; }
    .stError   { background: #fff5f5 !important; border: 1px solid #fed7d7 !important; border-radius: 8px; color: #c53030 !important; }
    .stInfo    { background: #ebf4ff !important; border: 1px solid #bee3f8 !important; border-radius: 8px; color: #2b6cb0 !important; }

    /* File uploader */
    [data-testid="stFileUploader"] {
        border: 2px dashed #cbd5e0 !important;
        border-radius: 14px !important;
        background: #f8f9fc !important;
        transition: border-color 0.2s, background-color 0.2s;
    }
    [data-testid="stFileUploader"]:hover { 
        border-color: #3182ce !important; 
        background: #ebf4ff !important;
    }

    /* Divider */
    hr { border-color: #e2e8f0 !important; margin: 2em 0; }

    /* Image caption */
    .stImage figcaption { color: #718096 !important; font-size: 0.8rem; }

    /* ── Chain of Title Timeline ── */
    .chain-timeline { position: relative; padding-left: clamp(1.5rem, 4vw, 2.5rem); margin: 1.5rem 0; }
    .chain-timeline::before {
        content: ''; position: absolute; left: 0.9rem; top: 0; bottom: 0;
        width: 3px; background: linear-gradient(180deg, #3182ce, #63b3ed);
        border-radius: 2px;
    }
    .chain-node {
        position: relative; margin-bottom: 1.5rem; padding: clamp(0.8rem, 2vw, 1.2rem);
        background: #ffffff; border: 1px solid #e2e8f0;
        border-radius: 10px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
    }
    .chain-node.current {
        border-color: #48bb78; background: #f0fff4;
        box-shadow: 0 0 12px rgba(72,187,120,0.15);
    }
    .chain-node.gap-warning {
        border-color: #ed8936; background: #fffaf0;
    }
    .chain-node .chain-date {
        display: inline-block; padding: 0.15rem 0.5rem; border-radius: 4px;
        background: #edf2f7; color: #4a5568; font-size: 0.75rem; font-weight: 600;
        margin-bottom: 0.4rem;
    }
    .chain-node .chain-party { color: #2d3748; font-size: 0.9rem; line-height: 1.4; }
    .gap-alert {
        padding: clamp(0.6rem, 2vw, 1rem); border-radius: 10px; text-align: center;
        background: #fff5f5; border: 1px dashed #fc8181;
        color: #c53030; font-size: 0.8rem; margin-bottom: 1.5rem;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────────────────
# Extraction prompt
# ─────────────────────────────────────────────────────────────────────────────
EXTRACTION_PROMPT = (
    "You are an expert Indian real estate data extractor. "
    "I am sending you ALL pages of an Indian property title deed. "
    "Return ONLY a valid JSON object (no markdown fences). "
    "IMPORTANT: Every field must be a nested object with exactly THREE keys: "
    "'value' (ENGLISH translation/transliteration of the extracted text), "
    "'value_hi' (the ORIGINAL Hindi/Devanagari text exactly as written in the document), "
    "and 'confidence' (strictly 'High', 'Medium', or 'Low').\n"
    "Use 'Low' if the handwriting is messy, smudged, or barely legible. "
    "Use 'Medium' if you can read it but are not 100% certain. "
    "Use 'High' only when the text is clearly readable.\n\n"
    "Return exactly these keys:\n"
    '{\n'
    '  "customer_name": {"value": "English name", "value_hi": "हिंदी नाम", "confidence": "High"},\n'
    '  "address": {"value": "English address", "value_hi": "हिंदी पता", "confidence": "High"},\n'
    '  "land_area": {"value": "area in English", "value_hi": "क्षेत्रफल", "confidence": "High"},\n'
    '  "dim_east": {"value": "", "value_hi": "", "confidence": "High"},\n'
    '  "dim_west": {"value": "", "value_hi": "", "confidence": "High"},\n'
    '  "dim_north": {"value": "", "value_hi": "", "confidence": "High"},\n'
    '  "dim_south": {"value": "", "value_hi": "", "confidence": "High"},\n'
    '  "bound_east": {"value": "", "value_hi": "", "confidence": "High"},\n'
    '  "bound_west": {"value": "", "value_hi": "", "confidence": "High"},\n'
    '  "bound_north": {"value": "", "value_hi": "", "confidence": "High"},\n'
    '  "bound_south": {"value": "", "value_hi": "", "confidence": "High"}\n'
    '}\n'
    "CRITICAL ACCURACY INSTRUCTIONS:\n"
    "1. For customer_name, extract the full name of the current buyer/applicant. Format the English value as: "
    "'Sh./Smt. [First Name] S/O or W/O or D/O [Father/Husband Name]'.\n"
    "2. For address, combine ALL location details into one complete address string.\n"
    "3. For land_area, include both the number and unit (e.g. '784.40 sq.ft').\n"
    "4. For dimensions (dim_east/west/etc.), extract only the dimension value.\n"
    "5. For boundaries (bound_east/west/etc.), extract the name of the neighbor/property on that side. Translate Hindi directions carefully: "
    "उत्तर = North, दक्षिण = South, पूर्व = East, पश्चिम = West.\n"
    "6. ACCURACY IS PARAMOUNT. Do NOT guess or infer — if you cannot read a value clearly, "
    "set value to 'Unclear', value_hi to 'अस्पष्ट', and confidence to 'Low'.\n"
    "7. If the document is in English only, set value_hi to the same text as value."
)

EXTRACTION_PROMPT_SITE_VISIT = """
You are a Senior Technical Valuer. Read this field runner's sketch/notes. The text may be in English or Hinglish. 
Extract the ACTUAL onsite details into a strict JSON object. If a value is missing or unreadable, return "N/A".
Return ONLY the raw JSON object, without markdown formatting.

Extract these exact keys:
"actual_north" -> Actual North boundary property/road
"actual_south" -> Actual South boundary property/road
"actual_east" -> Actual East boundary property/road
"actual_west" -> Actual West boundary property/road
"actual_land_area" -> Actual constructed/measured land area on site
"construction_stage" -> Stage of construction (e.g., Vacant, Under Construction, Fully Constructed)
"illegal_occupation_or_encroachment" -> true if there are mentions of illegal occupation, extensions outside bounds, or encroachments. False otherwise.
"site_remarks" -> Short string summarizing any risk notes, remarks, or observations from the valuer.
"""

EXTRACTION_PROMPT_FORMAT_CONVERTER = """
You are an expert Data Extractor. Read these scanned pages of a property valuation report. 
Extract the data and return a strict JSON object. If a value is missing, return "N/A". Return ONLY the raw JSON object, without markdown formatting.

Extract these exact keys:
"report_date" -> Date of report
"owner_name" -> Name of the property owner/borrower
"sale_deed_no" -> Sale deed number
"plot_no" -> Plot Number / Khasra No / Patta No
"road_width" -> Width of the road
"colony" -> Colony / Nagar / Sector
"landmark" -> Locality / Landmark
"city" -> Village / Town / City
"pincode" -> Pincode
"lat" -> Latitude
"lon" -> Longitude
"property_type" -> Type of Property (e.g. Residential, Commercial)
"land_level" -> Level of land with topographical conditions
"construction_observed" -> Any construction observed on plot
"civic_amenities" -> Civic Amenities like school, hospital etc
"transport_availability" -> Availability of local transport
"plot_area_doc" -> Plot Area as per documents (Sqft)
"plot_area_actual" -> Plot area as per actual site (Sqft)
"approved_built_up_area" -> Approved Built Up Area (in Sq.Ft.)
"north_boundary", "south_boundary", "east_boundary", "west_boundary" -> Boundaries as per actual site
"structure_type" -> Type of Structure
"occupancy" -> Occupancy Details (Self-Occupied / Rented / Vacant)
"current_life_years" -> Current Life of the structure in years
"projected_life_years" -> Projected Life of the Structure in years
"area_basement" -> Constructed area of Basement (in Sq.Ft., N/A if 0 or none)
"area_ground_floor" -> Constructed area of Ground Floor (in Sq.Ft., N/A if 0 or none)
"area_first_floor" -> Constructed area of First Floor (in Sq.Ft., N/A if 0 or none)
"area_second_floor" -> Constructed area of Second Floor (in Sq.Ft., N/A if 0 or none)
"area_third_floor" -> Constructed area of Third Floor (in Sq.Ft., N/A if 0 or none)
"land_rate" -> Rate per Sq.Ft for Land
"land_value" -> Amount in Rs for Land
"building_rate" -> Rate per Sq.Ft for Building
"building_value" -> Amount in Rs for Building
"total_market_value" -> Market value Total Valuation in numbers
"distress_value" -> Distressed / Forced Sale Value
"""

# ─────────────────────────────────────────────────────────────────────────────
# Bank template registry — add new banks here
# ─────────────────────────────────────────────────────────────────────────────
BANK_CONFIGS = {
    "Cholamandalam": {
        "template": "templates/chola_template.xlsx",
        "output_filename": "chola_valuation_report.xlsx",
        "cell_map": {
            "B5":  "customer_name",
            "B14": "address",
            "B15": "address",
            "B22": "land_area",
            "C22": "land_area",
            "B33": "dim_east",
            "B34": "dim_east",
            "C33": "dim_west",
            "C34": "dim_west",
            "D33": "dim_north",
            "D34": "dim_north",
            "E33": "dim_south",
            "E34": "dim_south",
            "B36": "bound_east",
            "B37": "bound_east",
            "C36": "bound_west",
            "C37": "bound_west",
            "D36": "bound_north",
            "D37": "bound_north",
            "E36": "bound_south",
            "E37": "bound_south",
        },
    },
    # ── Add more banks below ──────────────────────────────────────────────
    # "HDFC": {
    #     "template": "templates/hdfc_template.xlsx",
    #     "output_filename": "hdfc_valuation_report.xlsx",
    #     "cell_map": { ... },
    # },
}


# ─────────────────────────────────────────────────────────────────────────────
# Helper functions
# ─────────────────────────────────────────────────────────────────────────────

def pdf_to_pil_images(pdf_bytes: bytes, dpi: int = 200) -> list:
    """Render ALL pages of a PDF into PIL Images using PyMuPDF (fitz)."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    images = []
    zoom = dpi / 72
    mat = fitz.Matrix(zoom, zoom)
    for i in range(len(doc)):
        pixmap = doc[i].get_pixmap(matrix=mat, alpha=False)
        img = Image.frombytes("RGB", [pixmap.width, pixmap.height], pixmap.samples)
        images.append(img)
    doc.close()
    return images


def extract_with_gemini(api_key: str, images: list, model_name: str = "gemini-2.5-flash"):
    """
    Send ALL page images to Gemini Vision with the extraction prompt.
    Returns (parsed_dict, raw_text).
    """
    client = genai.Client(api_key=api_key)
    contents = [EXTRACTION_PROMPT] + images
    response = client.models.generate_content(
        model=model_name,
        contents=contents,
    )
    raw_text = response.text.strip()

    # Strip any accidental markdown fences
    clean = raw_text
    if clean.startswith("```"):
        clean = "\n".join(clean.split("\n")[1:])
    if clean.endswith("```"):
        clean = clean[: clean.rfind("```")]
    clean = clean.strip()

    try:
        return json.loads(clean), raw_text
    except json.JSONDecodeError as exc:
        raise ValueError(f"Gemini returned non-JSON output: {exc}") from exc


def extract_format_converter_with_gemini(api_key: str, images: list):
    """Sends the PIL images to Gemini and requests strictly formatted JSON for the Axis format."""
    client = genai.Client(api_key=api_key)
    payload = images + [EXTRACTION_PROMPT_FORMAT_CONVERTER]
    
    try:
        with st.spinner("🧠 Initializing Format Shifting Engine (Gemini 2.5 Pro)..."):
            response = client.models.generate_content(
                model="gemini-2.5-pro",
                contents=payload,
            )
    except Exception as e:
        if "429" in str(e) or "Quota exceeded" in str(e):
            st.warning("⚠️ Gemini 2.5 Pro rate limit reached. Automatically falling back to high-capacity Gemini 2.5 Flash model...")
            with st.spinner("⚡ Re-running extraction with Gemini 2.5 Flash..."):
                response = client.models.generate_content(
                    model="gemini-2.5-flash",
                    contents=payload,
                )
        else:
            raise e

    raw_text = response.text.strip()
    if raw_text.startswith("```json"):
        raw_text = raw_text[7:]
    if raw_text.endswith("```"):
        raw_text = raw_text[:-3]
        
    try:
        return json.loads(raw_text.strip()), raw_text
    except Exception as e:
        return {}, str(e)


def extract_site_visit_sketch(api_key: str, images: list):
    """Sends the PIL images to Gemini to extract ground truth from field runner handwritten sketches."""
    client = genai.Client(api_key=api_key)
    payload = images + [EXTRACTION_PROMPT_SITE_VISIT]
    
    try:
        response = client.models.generate_content(
            model="gemini-2.5-pro",
            contents=payload,
        )
    except Exception as e:
        if "429" in str(e) or "Quota exceeded" in str(e):
            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=payload,
            )
        else:
            raise e

    raw_text = response.text.strip()
    if raw_text.startswith("```json"):
        raw_text = raw_text[7:]
    if raw_text.endswith("```"):
        raw_text = raw_text[:-3]
        
    try:
        return json.loads(raw_text.strip()), raw_text
    except Exception as e:
        return {}, str(e)


from copy import copy

def inject_value_preserve_style(ws, coord, value):
    """Injects a value into a cell while strictly maintaining its original Excel formatting."""
    target_cell = ws[coord]
    
    # Cache the original styles before overwriting
    original_font = copy(target_cell.font) if target_cell.font else None
    original_border = copy(target_cell.border) if target_cell.border else None
    original_fill = copy(target_cell.fill) if target_cell.fill else None
    original_number_format = copy(target_cell.number_format) if target_cell.number_format else None
    original_protection = copy(target_cell.protection) if target_cell.protection else None
    original_alignment = copy(target_cell.alignment) if target_cell.alignment else None

    # Inject Value
    target_cell.value = value
    
    # Re-apply styles
    if original_font: target_cell.font = original_font
    if original_border: target_cell.border = original_border
    if original_fill: target_cell.fill = original_fill
    if original_number_format: target_cell.number_format = original_number_format
    if original_protection: target_cell.protection = original_protection
    if original_alignment: target_cell.alignment = original_alignment


def generate_axis_report(data: dict) -> io.BytesIO:
    """Loads the target template, maps the extracted JSON to specific cells, and returns a BytesIO buffer."""
    template_path = "templates/axis_template.xlsx"
    
    wb = openpyxl.load_workbook(template_path, data_only=False)
    
    if 'Sheet1' in wb.sheetnames:
        ws = wb['Sheet1']
    else:
        ws = wb.active
    
    inject_value_preserve_style(ws, 'D10', data.get('report_date', 'N/A'))
    inject_value_preserve_style(ws, 'D11', data.get('owner_name', 'N/A'))
    inject_value_preserve_style(ws, 'D22', data.get('sale_deed_no', 'N/A'))
    inject_value_preserve_style(ws, 'D23', data.get('plot_no', 'N/A'))
    inject_value_preserve_style(ws, 'J23', data.get('road_width', 'N/A'))
    inject_value_preserve_style(ws, 'D24', data.get('colony', 'N/A'))
    inject_value_preserve_style(ws, 'J24', data.get('landmark', 'N/A'))
    inject_value_preserve_style(ws, 'D25', data.get('city', 'N/A'))
    inject_value_preserve_style(ws, 'J26', data.get('pincode', 'N/A'))
    inject_value_preserve_style(ws, 'E28', data.get('lat', 'N/A'))
    inject_value_preserve_style(ws, 'K28', data.get('lon', 'N/A'))
    
    inject_value_preserve_style(ws, 'G31', data.get('property_type', 'N/A'))
    inject_value_preserve_style(ws, 'G32', data.get('land_level', 'N/A'))
    inject_value_preserve_style(ws, 'G33', data.get('construction_observed', 'N/A'))
    inject_value_preserve_style(ws, 'G37', data.get('civic_amenities', 'N/A'))
    inject_value_preserve_style(ws, 'G41', data.get('transport_availability', 'N/A'))
    
    inject_value_preserve_style(ws, 'E54', data.get('plot_area_doc', 'N/A'))
    inject_value_preserve_style(ws, 'K54', data.get('plot_area_actual', 'N/A'))
    
    inject_value_preserve_style(ws, 'H52', data.get('east_boundary', 'N/A'))
    inject_value_preserve_style(ws, 'H53', data.get('west_boundary', 'N/A'))
    inject_value_preserve_style(ws, 'H50', data.get('north_boundary', 'N/A'))
    inject_value_preserve_style(ws, 'H51', data.get('south_boundary', 'N/A'))
    
    inject_value_preserve_style(ws, 'B52', data.get('east_boundary', 'N/A'))
    inject_value_preserve_style(ws, 'B53', data.get('west_boundary', 'N/A'))
    inject_value_preserve_style(ws, 'B50', data.get('north_boundary', 'N/A'))
    inject_value_preserve_style(ws, 'B51', data.get('south_boundary', 'N/A'))
    
    inject_value_preserve_style(ws, 'K78', data.get('approved_built_up_area', 'N/A'))
    
    inject_value_preserve_style(ws, 'C83', data.get('area_basement', 'N/A'))
    inject_value_preserve_style(ws, 'C84', data.get('area_ground_floor', 'N/A'))
    inject_value_preserve_style(ws, 'C85', data.get('area_first_floor', 'N/A'))
    inject_value_preserve_style(ws, 'C86', data.get('area_second_floor', 'N/A'))
    inject_value_preserve_style(ws, 'C87', data.get('area_third_floor', 'N/A'))
    
    valid_floors = 0
    areas = [
        data.get('area_basement', 'N/A'),
        data.get('area_ground_floor', 'N/A'),
        data.get('area_first_floor', 'N/A'),
        data.get('area_second_floor', 'N/A'),
        data.get('area_third_floor', 'N/A')
    ]
    for a in areas:
        val = str(a).upper().strip()
        if val != 'N/A' and val != '0' and val != '':
            valid_floors += 1
            
    inject_value_preserve_style(ws, 'G62', valid_floors if valid_floors > 0 else 'N/A')
    
    inject_value_preserve_style(ws, 'G61', data.get('structure_type', 'N/A'))
    inject_value_preserve_style(ws, 'G63', data.get('occupancy', 'N/A'))
    inject_value_preserve_style(ws, 'E100', data.get('current_life_years', 'N/A'))
    inject_value_preserve_style(ws, 'K100', data.get('projected_life_years', 'N/A'))
    
    inject_value_preserve_style(ws, 'G107', data.get('land_rate', 'N/A'))
    inject_value_preserve_style(ws, 'J107', data.get('land_value', 'N/A'))
    inject_value_preserve_style(ws, 'G108', data.get('building_rate', 'N/A'))
    inject_value_preserve_style(ws, 'J108', data.get('building_value', 'N/A'))
    inject_value_preserve_style(ws, 'J121', data.get('total_market_value', 'N/A'))
    inject_value_preserve_style(ws, 'J122', data.get('distress_value', 'N/A'))
    
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    
    return output_stream


def get_val(data: dict, key: str, default: str = "N/A") -> str:
    """Safely extract 'value' from nested {value, value_hi, confidence} or plain string."""
    field = data.get(key, default)
    if isinstance(field, dict):
        return field.get("value", default)
    return field if field else default


def get_hindi(data: dict, key: str, default: str = "—") -> str:
    """Extract the original Hindi text from nested dict."""
    field = data.get(key, default)
    if isinstance(field, dict):
        return field.get("value_hi", default)
    return default


def get_conf(data: dict, key: str) -> str:
    """Get confidence level from nested dict. Defaults to 'High'."""
    field = data.get(key)
    if isinstance(field, dict):
        return field.get("confidence", "High")
    return "High"


def conf_badge(level: str) -> str:
    """Return colored emoji badge for confidence level."""
    badges = {"High": "🟢 High", "Medium": "🟡 Medium", "Low": "🔴 Low"}
    return badges.get(level, "⚪ Unknown")


def sanitize_filename(name: str) -> str:
    """Sanitize a string for use as a filename."""
    clean = re.sub(r'[^\w\s-]', '', name).strip()
    clean = re.sub(r'[\s]+', '_', clean)
    return clean[:50] if clean else "report"


# ─────────────────────────────────────────────────────────────────────────────
# Satellite Discrepancy Engine
# ─────────────────────────────────────────────────────────────────────────────
# Prompts
# ─────────────────────────────────────────────────────────────────────────────
DISCREPANCY_PROMPT = (
    "You are a PropTech Risk Officer. I am giving you:\n"
    "1. The extracted details of a property deed (JSON below)\n"
    "2. Up to three current satellite images of the property coordinates from different providers (e.g., Google, Esri, Mapbox).\n\n"
    "DEED DATA:\n{deed_json}\n\n"
    "Compare the deed data against the satellite images and produce a DISCREPANCY REPORT. Analyze carefully:\n"
    "- Does the deed describe empty/vacant land while the satellite shows a built structure (or vice versa)?\n"
    "- Are there obvious boundary discrepancies or encroachments visible?\n"
    "- Any signs of unauthorized construction, environmental risk, or flood-prone terrain?\n"
    "- Does the visible plot size roughly match the deed's land area?\n"
    "- Note any differences between the satellite images (e.g. one might show newer construction than another).\n\n"
    "Output EXACTLY in this format:\n"
    "RISK LEVEL: [HIGH / MEDIUM / LOW / NONE]\n"
    "FINDINGS:\n"
    "- [finding 1]\n"
    "- [finding 2]\n"
    "- ...\n"
    "RECOMMENDATION: [one-line action item for the bank valuer]"
)

INSIGHTS_PROMPT = (
    "You are a Real Estate Valuer and Geography Expert. I am giving you up to three satellite images "
    "of a specific property coordinate from different providers (e.g., Google, Esri, Mapbox). "
    "There is NO deed data available. Provide a pure visual analysis of the property and its surroundings.\n\n"
    "Analyze thoroughly:\n"
    "- Land classification (residential, commercial, agricultural, industrial, barren, etc.)\n"
    "- Surrounding infrastructure (proximity to major roads, highways, water bodies, or urban density)\n"
    "- Development density (highly developed area vs. open greenfield)\n"
    "- Potential risks (flood plains, dense forests, steep terrain, industrial hazards, proximity to coastal/river banks)\n\n"
    "Output EXACTLY in this format:\n"
    "PROPERTY TYPE: [Likely classification]\n"
    "DEVELOPMENT LEVEL: [High / Medium / Low / Undeveloped]\n"
    "GEOGRAPHIC INSIGHTS:\n"
    "- [insight 1]\n"
    "- [insight 2]\n"
    "- ...\n"
    "NOTABLE RISKS: [List any visible environmental or locational risks, or 'None visible']"
)


import math

def deg2num(lat_deg: float, lon_deg: float, zoom: int) -> tuple:
    """Calculate the Slippy Map tile X and Y from Lat/Lon."""
    lat_rad = math.radians(lat_deg)
    n = 2.0 ** zoom
    xtile = int((lon_deg + 180.0) / 360.0 * n)
    ytile = int((1.0 - math.asinh(math.tan(lat_rad)) / math.pi) / 2.0 * n)
    return (xtile, ytile)


from PIL.ExifTags import TAGS, GPSTAGS

def get_decimal_from_dms(dms, ref):
    """Convert EXIF GPS DMS (Degrees, Minutes, Seconds) to Decimal Degrees."""
    degrees = float(dms[0])
    minutes = float(dms[1])
    seconds = float(dms[2])
    
    decimal = degrees + (minutes / 60.0) + (seconds / 3600.0)
    if ref in ['S', 'W']:
        decimal = -decimal
    return decimal

def get_exif_gps_coords(image: Image.Image) -> tuple:
    """Extract Latitude and Longitude from PIL Image EXIF data."""
    exif_data = image._getexif()
    if not exif_data:
        return None
    
    gps_info = None
    for tag_id, value in exif_data.items():
        tag_name = TAGS.get(tag_id, tag_id)
        if tag_name == "GPSInfo":
            gps_info = {GPSTAGS.get(t, t): gps_data for t, gps_data in value.items()}
            break
            
    if not gps_info:
        return None
        
    try:
        lat = get_decimal_from_dms(gps_info['GPSLatitude'], gps_info['GPSLatitudeRef'])
        lon = get_decimal_from_dms(gps_info['GPSLongitude'], gps_info['GPSLongitudeRef'])
        return lat, lon
    except KeyError:
        return None


def download_google_satellite(lat: float, lon: float, api_key: str, zoom: int = 20) -> Image.Image:
    url = f"https://maps.googleapis.com/maps/api/staticmap?center={lat},{lon}&zoom={zoom}&size=600x600&maptype=satellite&key={api_key}"
    resp = requests.get(url, timeout=10)
    resp.raise_for_status()
    return Image.open(io.BytesIO(resp.content))


def download_esri_satellite(lat: float, lon: float, zoom: int = 18) -> Image.Image:
    x, y = deg2num(lat, lon, zoom)
    url = f"https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{zoom}/{y}/{x}"
    resp = requests.get(url, timeout=10)
    resp.raise_for_status()
    return Image.open(io.BytesIO(resp.content))


def download_mapbox_satellite(lat: float, lon: float, api_key: str, zoom: int = 18) -> Image.Image:
    url = f"https://api.mapbox.com/styles/v1/mapbox/satellite-v9/static/{lon},{lat},{zoom},0/600x600?access_token={api_key}"
    resp = requests.get(url, timeout=10)
    resp.raise_for_status()
    return Image.open(io.BytesIO(resp.content))


def get_all_satellite_imagery(lat: float, lon: float) -> dict:
    """Fetch available satellite images from all configured providers."""
    images = {}
    
    # 1. Esri (Free, no key needed)
    try:
        images["Esri World Imagery"] = download_esri_satellite(lat, lon, zoom=19)
    except Exception as e:
        images["Esri World Imagery"] = f"Error: {e}"

    # 2. Google Maps
    if GMAPS_API_KEY:
        try:
            images["Google Maps Satellite"] = download_google_satellite(lat, lon, GMAPS_API_KEY, zoom=20)
        except Exception as e:
            images["Google Maps Satellite"] = f"Error: {e}"
            
    # 3. Mapbox
    if MAPBOX_API_KEY:
        try:
            images["Mapbox Satellite"] = download_mapbox_satellite(lat, lon, MAPBOX_API_KEY, zoom=19)
        except Exception as e:
            images["Mapbox Satellite"] = f"Error: {e}"
            
    return images


def run_multi_source_analysis(api_key: str, images_dict: dict, deed_data: dict = None) -> str:
    """Send multiple satellite images + (optional) deed JSON to Gemini for risk/insight analysis."""
    client = genai.Client(api_key=api_key)
    
    # Filter out failures so we only send valid PIL images
    valid_images = [img for name, img in images_dict.items() if isinstance(img, Image.Image)]
    
    if deed_data:
        # Discrepancy Check Mode
        clean_data = {}
        for key in ["customer_name", "address", "land_area",
                    "dim_east", "dim_west", "dim_north", "dim_south",
                    "bound_east", "bound_west", "bound_north", "bound_south"]:
            clean_data[key] = get_val(deed_data, key)
        prompt = DISCREPANCY_PROMPT.format(deed_json=json.dumps(clean_data, indent=2))
        prompt_content = [prompt] + valid_images
    else:
        # Standalone Insights Mode
        prompt_content = [INSIGHTS_PROMPT] + valid_images

    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=prompt_content,
    )
    return response.text.strip()


def parse_risk_level(report: str) -> str:
    """Extract risk level from the discrepancy report."""
    report_upper = report.upper()
    if "RISK LEVEL: HIGH" in report_upper or "RISK LEVEL:HIGH" in report_upper:
        return "HIGH"
    elif "RISK LEVEL: MEDIUM" in report_upper or "RISK LEVEL:MEDIUM" in report_upper:
        return "MEDIUM"
    elif "RISK LEVEL: LOW" in report_upper or "RISK LEVEL:LOW" in report_upper:
        return "LOW"
    return "NONE"


def generate_bank_report(bank_key: str, extracted_data: dict) -> io.BytesIO:
    """Generate Excel report for the selected bank using its config from BANK_CONFIGS."""
    config = BANK_CONFIGS[bank_key]
    wb = openpyxl.load_workbook(config["template"])
    ws = wb.active

    for cell_ref, data_key in config["cell_map"].items():
        ws[cell_ref] = get_val(extracted_data, data_key)

    out_stream = io.BytesIO()
    wb.save(out_stream)
    out_stream.seek(0)
    return out_stream


# ─────────────────────────────────────────────────────────────────────────────
# Display configuration — categories, order, and labels
# ─────────────────────────────────────────────────────────────────────────────
CATEGORY_CONFIG = [
    ("Buyer",    "👤 Buyer / Purchaser (क्रेता)",  [
        ("Name", "Name"), ("Age", "Age"), ("Address", "Address"),
    ]),
    ("Seller",   "👤 Seller / Vendor (विक्रेता)",   [
        ("Name", "Name"), ("Age", "Age"), ("Address", "Address"),
    ]),
    ("Property_Identifiers", "📍 Property Identifiers", None),  # special rendering
    ("Total_Area", "📐 Total Area", [
        ("Area_Sq_Meters", "Area (sq. meters)"), ("Area_Sq_Feet", "Area (sq. feet)"), ("Construction_Details", "Construction Details"),
    ]),
    ("Transaction", "💰 Transaction / Stamp Details", [
        ("Sale_Price", "Sale Price"), ("Stamp_Duty", "Stamp Duty"),
    ]),
    ("Registration", "📋 Registration Details", [
        ("Registration_Date", "Registration Date"), ("Registration_Number", "Registration Number"), ("Sub_Registrar_Office", "Sub-Registrar Office"),
    ]),
]

BOUNDARY_DIRECTIONS = ["North", "South", "East", "West"]
BOUNDARY_HINDI = {"North": "उत्तर", "South": "दक्षिण", "East": "पूर्व", "West": "पश्चिम"}


def render_category_table(st_container, data: dict, fields: list):
    """Render a category as a styled table."""
    import pandas as pd
    rows = []
    for key, label in fields:
        val = data.get(key, "N/A")
        if isinstance(val, dict):
            val = ", ".join(f"{k}: {v}" for k, v in val.items())
def render_boundaries_table(st_container, boundaries: dict):
    pass

def render_witnesses(st_container, witnesses):
    pass


# ─────────────────────────────────────────────────────────────────────────────
# Chain of Title — Helper functions
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_name(name: str) -> str:
    """Normalize a name for fuzzy matching: lowercase, strip honorifics."""
    import re
    n = name.lower().strip()
    # Remove common prefixes and relation markers
    for prefix in ["sh.", "shri", "smt.", "smt", "sh", "mr.", "mrs.", "ms."]:
        n = n.replace(prefix, "")
    # Remove relation markers and everything after
    for marker in [" s/o ", " w/o ", " d/o ", " son of ", " wife of ", " daughter of "]:
        if marker in n:
            n = n.split(marker)[0]
    return re.sub(r"\s+", " ", n).strip()


def _name_match(name_a: str, name_b: str) -> bool:
    """Fuzzy check if two names refer to the same person."""
    a = _normalize_name(name_a)
    b = _normalize_name(name_b)
    if not a or not b:
        return False
    # Exact or substring match
    if a == b or a in b or b in a:
        return True
    # Token overlap: at least 2 tokens in common
    tokens_a = set(a.split())
    tokens_b = set(b.split())
    overlap = tokens_a & tokens_b
    return len(overlap) >= 2


def _parse_date_for_sort(date_str: str):
    """Try to parse a date string for sorting. Returns a sortable tuple."""
    import re
    if not date_str or date_str == "N/A":
        return (9999, 99, 99)
    # Try DD-MM-YYYY or DD/MM/YYYY
    m = re.match(r"(\d{1,2})[\-/](\d{1,2})[\-/](\d{4})", date_str)
    if m:
        return (int(m.group(3)), int(m.group(2)), int(m.group(1)))
    # Try YYYY-MM-DD
    m = re.match(r"(\d{4})[\-/](\d{1,2})[\-/](\d{1,2})", date_str)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    # Just extract year
    m = re.search(r"(\d{4})", date_str)
    if m:
        return (int(m.group(1)), 1, 1)
    return (9999, 99, 99)


def build_ownership_chain(deed_list: list) -> list:
    """
    Given a list of extracted deed dicts, build a sequence.
    Sorting assumes documents are uploaded in order since dates are removed.
    Returns a list of dicts with added 'chain_status' and 'gap_note' fields.
    """
    chain = []
    for i, deed in enumerate(deed_list):
        entry = {
            "data": deed,
            "date_str": "N/A", # Dates were removed in new prompt format
            "is_current": (i == len(deed_list) - 1),
            "has_gap": False,
            "gap_note": "",
        }
        
        # Check chain continuity: without explicit sellers, we assume valid sequence 
        # or gap if names change unexpectedly. For now, mark all as potential gaps
        # because we don't have Seller name to confirm.
        if i > 0:
            prev_buyer = deed_list[i - 1].get("customer_name", "N/A")
            this_buyer = deed.get("customer_name", "N/A")
            if prev_buyer and this_buyer and prev_buyer != this_buyer:
                entry["has_gap"] = True
                entry["gap_note"] = f"Previous buyer ({prev_buyer}) -> New applicant ({this_buyer}). Missing seller link."

        chain.append(entry)

    return chain


def render_chain_timeline(chain: list):
    """Render the ownership chain as a visual timeline."""
    html = '<div class="chain-timeline">'

    for entry in chain:
        deed = entry["data"]
        buyer = get_val(deed, "customer_name")
        seller = "N/A (Hidden in new format)"
        prop_type = ""
        reg_no = ""
        sale_price = "N/A"

        css_class = "chain-node"
        badges = ""
        if entry["is_current"]:
            css_class += " current"
            badges += '<span class="node-badge badge-current">★ Current Deed</span>'
        if entry["has_gap"]:
            css_class += " gap-warning"
            badges += '<span class="node-badge badge-gap">⚠ Gap</span>'

        html += f'''
        <div class="{css_class}">
            <div class="node-date">{entry["date_str"]}{badges}</div>
            <div class="node-title">{prop_type or "Property Transfer"}</div>
            <div class="node-detail">
                <strong>Seller:</strong> {seller}<br>
                <strong>Buyer:</strong> {buyer}<br>
                <strong>Sale Price:</strong> {sale_price}
                {f"<br><strong>Reg. No:</strong> {reg_no}" if reg_no and reg_no != 'N/A' else ""}
            </div>
            {f'<div style="margin-top:0.5rem; font-size:0.78rem; color:#ffa726;">⚠ {entry["gap_note"]}</div>' if entry["has_gap"] else ""}
        </div>
        '''

    html += '</div>'

    # Chain status summary
    gaps = sum(1 for e in chain if e["has_gap"])
    if gaps == 0:
        html += '<div class="chain-status chain-complete">✅ Chain is COMPLETE — no gaps detected</div>'
    else:
        html += f'<div class="chain-status chain-broken">⚠️ {gaps} gap{"s" if gaps > 1 else ""} detected in the ownership chain</div>'

    return html


# ─────────────────────────────────────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📄 docfinder")
    st.markdown("**Powered by Gemini 2.5 Flash**")

    # API status indicator (no input needed from user)
    if GEMINI_API_KEY:
        st.success("AI Engine: Connected ✓", icon="🤖")
    else:
        st.error("AI Engine: Not configured", icon="🚨")
        st.caption("Admin: add `GEMINI_API_KEY` to `.env` file")


    if GMAPS_API_KEY:
        st.success("Maps Engine: Connected ✓", icon="🛰️")
    else:
        st.error("Maps Engine: Not configured", icon="🛰️")
        st.caption("Admin: add `GMAPS_API_KEY` to `.env` file")

    if MAPBOX_API_KEY:
        st.success("Mapbox Engine: Connected ✓", icon="🌍")
    else:
        st.error("Mapbox Engine: Not configured", icon="🌍")
        st.caption("Admin: add `MAPBOX_API_KEY` to `.env` file")


    st.markdown(
        """
        **Modes**

        📄 **Single Deed** — Upload 1 PDF, extract data
        
        🌍 **Property Insights** — Enter coordinates, get AI analysis (No Deed)

        🔗 **Chain of Title** — Upload multiple deeds,
        trace ownership from original to current owner
        
        🔄 **Format Converter** — Bulk structural migration to Axis Bank 2026 format

        ---

        **Fields extracted per deed**
        - Buyer & Seller (name, age, address)
        - Property Identifiers (patta/plot/khasra no.)
        - Area, Boundaries, Transaction
        - Registration & Witnesses

        ---
        *Handles printed and handwritten Hindi*
        """
    )

    st.caption(f"Model: `{DEFAULT_MODEL}` · Antigravity PropTech v3.0")


# ─────────────────────────────────────────────────────────────────────────────
# Main header
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    "<h1 style='text-align:center;'>docfinder</h1>",
    unsafe_allow_html=True,
)
st.markdown(
    "<p style='text-align:center; color:#7eb8f7; margin-top:-0.5rem;'>"
    "Upload any Indian property deed · AI reads every page · Structured data in seconds"
    "</p>",
    unsafe_allow_html=True,
)
mode = st.radio(
    "Choose mode",
    ["📄 Single Deed", "🌍 Property Insights (No Deed)", "📦 Batch Processing", "🔗 Chain of Title", "🔄 Format Converter"],
    horizontal=True,
    label_visibility="collapsed",
)
# ═══════════════════════════════════════════════════════════════════════════════
# MODE 1: SINGLE DEED (original behavior)
# ═══════════════════════════════════════════════════════════════════════════════
if mode == "📄 Single Deed":

    st.markdown('<div class="avm-card">', unsafe_allow_html=True)
    st.markdown('<span class="badge badge-blue">Step 1</span>', unsafe_allow_html=True)
    st.markdown("### 📄 Upload Property Document")
    st.markdown("Upload a scanned property deed PDF — any number of pages, printed or handwritten Hindi.")

    deed_file = st.file_uploader(
        "Step 1: Upload Property Document (Mandatory)",
        type=["pdf"],
        key="single_upload",
    )
    
    site_file = st.file_uploader(
        "Step 2: Upload Site Visit Sheet / Sketch (Optional)",
        type=["pdf", "png", "jpg", "jpeg"],
        key="site_upload",
    )
    st.markdown("</div>", unsafe_allow_html=True)

    if deed_file is not None:

        col_img, col_results = st.columns([1, 1.4], gap="large")

        with col_img:
            st.markdown('<div class="avm-card">', unsafe_allow_html=True)
            st.markdown('<span class="badge badge-blue">Step 2</span>', unsafe_allow_html=True)
            st.markdown("### 🖼️ Document Preview")

            with st.spinner("Rendering all pages…"):
                try:
                    pdf_bytes = deed_file.read()
                    page_images = pdf_to_pil_images(pdf_bytes, dpi=300)
                    num_pages = len(page_images)

                    st.image(
                        page_images[0],
                        use_container_width=True,
                        caption=f"Page 1 of {num_pages} · All pages will be analyzed"
                    )
                    if num_pages > 1:
                        st.info(f"📄 {num_pages} pages detected — AI will read all of them.", icon="📄")
                    render_ok = True
                except Exception as exc:
                    st.error(f"Failed to render PDF: {exc}")
                    render_ok = False
                    
            if site_file is not None:
                st.markdown("---")
                st.markdown("#### Site Sketch Preview")
                try:
                    if site_file.name.lower().endswith('.pdf'):
                        site_bytes = site_file.read()
                        site_images = pdf_to_pil_images(site_bytes, dpi=200)
                    else:
                        site_images = [Image.open(site_file)]
                    st.image(site_images[0], use_container_width=True, caption="Site Visit Sketch")
                    site_render_ok = True
                except Exception as exc:
                    st.error(f"Failed to render site sketch: {exc}")
                    site_render_ok = False
            else:
                site_render_ok = False
                site_images = []

            st.markdown("</div>", unsafe_allow_html=True)

        with col_results:
            st.markdown('<div class="avm-card">', unsafe_allow_html=True)
            st.markdown('<span class="badge badge-amber">Step 3</span>', unsafe_allow_html=True)
            st.markdown("### 🤖 AI Extraction")

            if not GEMINI_API_KEY:
                st.error("AI Engine not configured. Please contact the administrator.", icon="🚨")
            elif not render_ok:
                st.error("Cannot extract — PDF rendering failed in Step 2.")
            else:
                if st.button("⚡ Run AI Analysis", use_container_width=True, key="single_extract", type="primary"):
                    st.session_state["extracted_data"] = None
                    st.session_state["site_extracted_data"] = None
                    
                    spinner_msg = f"Analyzing {num_pages} page{'s' if num_pages > 1 else ''} with Gemini Vision…"
                    with st.spinner(spinner_msg):
                        try:
                            extracted_data, raw_response = extract_with_gemini(
                                GEMINI_API_KEY, page_images, DEFAULT_MODEL
                            )
                            st.session_state["extracted_data"] = extracted_data
                            st.session_state["raw_response"]   = raw_response
                            st.success(f"Deed Extraction complete! ({num_pages} pages analyzed)", icon="✅")
                        except ValueError as exc:
                            st.error(str(exc), icon="⚠️")
                            st.session_state["extracted_data"] = None
                            st.session_state["raw_response"]   = str(exc)
                        except Exception as exc:
                            st.error(f"AI Engine error: {exc}", icon="🚨")
                            with st.expander("Technical details"):
                                st.code(traceback.format_exc())
                            st.session_state["extracted_data"] = None
                            st.session_state["raw_response"]   = None
                            
                    if site_render_ok and site_images:
                        with st.spinner("Analyzing handwritten Site Sketch…"):
                            try:
                                site_extracted, site_raw = extract_site_visit_sketch(
                                    GEMINI_API_KEY, site_images
                                )
                                st.session_state["site_extracted_data"] = site_extracted
                                st.success("Site Sketch Extraction complete!", icon="✅")
                            except Exception as exc:
                                st.error(f"Failed to extract site sketch: {exc}")

            st.markdown("</div>", unsafe_allow_html=True)

        # ── Results display ──────────────────────────────────────────────────
        if st.session_state.get("extracted_data"):
            extracted = st.session_state["extracted_data"]
            site_extracted = st.session_state.get("site_extracted_data")

        
            st.markdown('<span class="badge badge-green">Results</span>', unsafe_allow_html=True)
            st.markdown("## 📊 Extracted Property Data")


            st.markdown("### ✨ Key Property Metrics")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric(label=f"👤 Borrower {conf_badge(get_conf(extracted, 'customer_name'))}", value=get_val(extracted, "customer_name") or "N/A")
            with col2:
                st.metric(label=f"📏 Total Land Area {conf_badge(get_conf(extracted, 'land_area'))}", value=get_val(extracted, "land_area") or "N/A")
            with col3:
                st.metric(label="🏠 Property Type", value="N/A")
            with col4:
                st.metric(label="📅 Date of Visit", value="N/A")

            st.markdown("---")

            col_add, col_bound = st.columns([2, 1])
            
            with col_add:
                st.markdown(f"### 📍 Property Address {conf_badge(get_conf(extracted, 'address'))}")
                st.info(get_val(extracted, "address") or "No address extracted.")
                
            with col_bound:
                st.markdown("### 🧭 Boundaries")
                st.markdown(f"""
                - **North:** {get_val(extracted, 'bound_north') or 'N/A'} {conf_badge(get_conf(extracted, 'bound_north'))}
                - **South:** {get_val(extracted, 'bound_south') or 'N/A'} {conf_badge(get_conf(extracted, 'bound_south'))}
                - **East:** {get_val(extracted, 'bound_east') or 'N/A'} {conf_badge(get_conf(extracted, 'bound_east'))}
                - **West:** {get_val(extracted, 'bound_west') or 'N/A'} {conf_badge(get_conf(extracted, 'bound_west'))}
                """)
                
            # ── TRUTH ENGINE: Site Verification ──────────────────────────────
            if site_extracted:
                st.markdown("---")
                st.markdown('<span class="badge badge-red">Truth Engine</span>', unsafe_allow_html=True)
                st.markdown("### ⚠️ AI Risk & Discrepancy Insights")
                
                legal_area = get_val(extracted, "land_area")
                actual_area = site_extracted.get("actual_land_area", "N/A")
                encroachment = site_extracted.get("illegal_occupation_or_encroachment", False)
                
                col_te1, col_te2 = st.columns(2)
                with col_te1:
                    st.metric("Deed Area (Legal)", legal_area or "N/A")
                with col_te2:
                    st.metric("Site Sketch Area (Actual)", actual_area)
                
                st.markdown("#### Discrepancy Engine Flags")
                # Very basic string comparison for demo logic. Real life would use NLP/regex conversion of units.
                if str(legal_area).lower().replace(" ","") != str(actual_area).lower().replace(" ","") and actual_area != "N/A" and legal_area:
                     st.warning(f"⚠️ **Area Mismatch:** Legal deed states {legal_area} but site inspection notes {actual_area}.")
                
                if encroachment is True or str(encroachment).lower() == 'true':
                     st.error("🚨 **Encroachment / Illegal Occupation Flag:** The field runner noted unauthorized construction or occupation on the property.")
                elif encroachment is False or str(encroachment).lower() == 'false':
                     st.success("✅ **Clear Title:** No illegal occupation or encroachments noted onsite.")
                     
                remarks = site_extracted.get("site_remarks", "")
                if remarks and remarks != "N/A":
                     st.info(f"**Field Runner Remarks:** {remarks}")

            # Download button is now outside of tabs and inside a clean full-width container
            st.markdown("---")
            st.markdown("### 📥 Download Report")
            
            col_bank, _ = st.columns([1, 1])
            with col_bank:
                bank_names = list(BANK_CONFIGS.keys())
                selected_bank = st.selectbox(
                    "Select Bank Template",
                    bank_names,
                    key="bank_selector",
                )
            
            download_container = st.container()
            with download_container:
                try:
                    config = BANK_CONFIGS[selected_bank]
                    # Pass both the deed payload and the optional site payload to the injector
                    excel_data = generate_bank_report(selected_bank, extracted, site_data=site_extracted)
                    st.download_button(
                        label=f"⬇️ Download {selected_bank} Report (.xlsx)",
                        data=excel_data,
                        file_name=config["output_filename"],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Failed to generate Excel report: {e}")

            # ── SATELLITE RISK ANALYSIS ───────────────────────────────────────
        
            st.markdown('<span class="badge badge-red">Risk Engine</span>', unsafe_allow_html=True)
            st.markdown("## 🛰️ Risk Analysis: Satellite vs. Deed")

            st.markdown('<div class="avm-card">', unsafe_allow_html=True)
            st.markdown("### 📍 Property Coordinates")
            st.markdown("Enter the latitude and longitude of the property to view satellite imagery and run risk analysis.")

            coord_col1, coord_col2 = st.columns(2)
            with coord_col1:
                sat_lat = st.number_input("Latitude", value=26.9124, format="%.6f", key="single_lat")
            with coord_col2:
                sat_lon = st.number_input("Longitude", value=75.7873, format="%.6f", key="single_lon")
            st.markdown('</div>', unsafe_allow_html=True)

            map_col, action_col = st.columns([1.5, 1])

            with map_col:
                st.markdown('<div class="avm-card">', unsafe_allow_html=True)
                st.markdown("### 🗺️ Satellite Map Preview")
                m = folium.Map(location=[sat_lat, sat_lon], zoom_start=18)
                # Satellite imagery base layer
                folium.TileLayer(
                    tiles="https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
                    attr="Esri", name="Satellite", overlay=False,
                ).add_to(m)
                # Labels overlay (roads, areas, landmarks)
                folium.TileLayer(
                    tiles="https://server.arcgisonline.com/ArcGIS/rest/services/Reference/World_Boundaries_and_Places/MapServer/tile/{z}/{y}/{x}",
                    attr="Esri", name="Labels", overlay=True,
                ).add_to(m)
                folium.Marker(
                    [sat_lat, sat_lon],
                    popup=f"Property: {get_val(extracted, 'address')}",
                    icon=folium.Icon(color="red", icon="home"),
                ).add_to(m)
                st_folium(m, width=700, height=400, key="single_map")
                st.markdown('</div>', unsafe_allow_html=True)

            with action_col:
                st.markdown('<div class="avm-card">', unsafe_allow_html=True)
                st.markdown("### 🔍 AI Discrepancy Check")
                st.markdown(
                    "Downloads a high-res satellite image from Google Maps "
                    "and asks Gemini to compare it against the deed data."
                )

                if st.button("🔍 Run AI Discrepancy Check", use_container_width=True, key="single_risk_check", type="primary"):
                    with st.spinner("📡 Downloading satellite imagery from multiple sources…"):
                        try:
                            images_dict = get_all_satellite_imagery(sat_lat, sat_lon)
                            st.session_state["sat_images_dict"] = images_dict
                        except Exception as exc:
                            st.error(f"Failed to fetch imagery: {exc}")
                            images_dict = None

                    if images_dict:
                        # Display images in a grid
                        cols = st.columns(len(images_dict))
                        for i, (provider, img) in enumerate(images_dict.items()):
                            with cols[i]:
                                if isinstance(img, Image.Image):
                                    st.image(img, caption=provider, use_container_width=True)
                                else:
                                    st.error(f"{provider}\n{img}", icon="❌")

                        with st.spinner("🤖 Running AI discrepancy analysis…"):
                            try:
                                report = run_multi_source_analysis(GEMINI_API_KEY, images_dict, extracted)
                                st.session_state["risk_report"] = report
                            except Exception as exc:
                                st.error(f"AI analysis failed: {exc}")
                                with st.expander("Details"):
                                    st.code(traceback.format_exc())

                st.markdown('</div>', unsafe_allow_html=True)

            # ── Discrepancy Report ────────────────────────────────────────────
            if st.session_state.get("risk_report"):
                report = st.session_state["risk_report"]
                risk_level = parse_risk_level(report)

            
                st.markdown('<span class="badge badge-red">Report</span>', unsafe_allow_html=True)
                st.markdown("## 📋 Discrepancy Report")

                st.markdown('<div class="avm-card">', unsafe_allow_html=True)
                if risk_level == "HIGH":
                    st.error("🚨 RISK LEVEL: HIGH — Significant discrepancies detected", icon="🚨")
                elif risk_level == "MEDIUM":
                    st.warning("⚠️ RISK LEVEL: MEDIUM — Some concerns require verification", icon="⚠️")
                else:
                    st.info(f"✅ RISK LEVEL: {risk_level} — No significant discrepancies", icon="✅")
                st.markdown(report)
                st.markdown('</div>', unsafe_allow_html=True)

                if st.session_state.get("sat_img"):
                    with st.expander("🛰️ Satellite Image Used for Analysis"):
                        st.image(st.session_state["sat_img"], caption="Google Maps Static API — Satellite View")

    else:
        # Idle hero state

        st.markdown(
            """
            <div style="text-align:center; padding: 3rem 1rem;">
                <h2 style="color:#2d3748; margin-bottom:0.5rem;">Ready to Extract</h2>
                <p style="color:#607d8b; max-width:480px; margin:0 auto;">
                    Upload a scanned Indian property document PDF above.
                    The AI will read every page and extract Buyer, Seller,
                    Property details, Area, Boundaries, Transaction value,
                    and Registration information — even from handwritten Hindi.
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )


        feat1, feat2, feat3, feat4 = st.columns(4)
        for col, icon, title, desc in [
            (feat1, "🔍", "Vision AI",      "Gemini 2.5 Flash reads printed & handwritten Hindi"),
            (feat2, "📄", "Multi-Page",     "Reads every page of the document, not just page 1"),
            (feat3, "🧭", "Full Extraction", "Buyer, Seller, Area, Boundaries, Transaction, Registration"),
            (feat4, "⚡", "Fast & Accurate", "Structured JSON in seconds, translated to English"),
        ]:
            with col:
                st.markdown(
                    f"""
                    <div class="avm-card" style="text-align:center; padding:1.2rem;">
                        <div style="font-size:2rem;">{icon}</div>
                        <div style="font-weight:600; color:#7eb8f7; margin:0.4rem 0 0.2rem;">{title}</div>
                        <div style="font-size:0.82rem; color:#7a90a8;">{desc}</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )


# ═══════════════════════════════════════════════════════════════════════════════
# MODE 4: PROPERTY INSIGHTS (NO DEED)
# ═══════════════════════════════════════════════════════════════════════════════
elif mode == "🌍 Property Insights (No Deed)":

    st.markdown('<div class="avm-card">', unsafe_allow_html=True)
    st.markdown('<span class="badge badge-amber">Earth Engine</span>', unsafe_allow_html=True)
    st.markdown("### 🌍 AI Property Insights")
    st.markdown(
        "Upload a ground-level property photo or enter coordinates to analyze the location. "
        "Gemini will assess land type, development density, infrastructure, and potential risks "
        "using up to 3 different high-res satellite providers simultaneously."
    )

    # Image Upload for EXIF extraction
    ground_img_file = st.file_uploader(
        "📸 Upload a Property Photo (Optional, extracts GPS)",
        type=["jpg", "jpeg", "png"],
        key="insights_upload"
    )

    ground_image = None
    if ground_img_file:
        try:
            ground_image = Image.open(ground_img_file)
            st.image(ground_image, caption="Uploaded Property Photo", use_container_width=True)
            
            # Try to extract GPS from EXIF
            coords = get_exif_gps_coords(ground_image)
            if coords:
                lat, lon = coords
                
                # Auto-fill only once per file upload to allow user overrides later
                file_hash = f"{ground_img_file.name}_{ground_img_file.size}"
                if st.session_state.get("last_uploaded_file_hash") != file_hash:
                    st.session_state["last_uploaded_file_hash"] = file_hash
                    st.session_state["insights_lat"] = float(lat)
                    st.session_state["insights_lon"] = float(lon)
                    st.rerun()

                st.success(f"📍 Found GPS coordinates in image: {lat:.6f}, {lon:.6f}", icon="✅")
            else:
                st.info("No GPS data found in the image. Please enter coordinates manually below.")
                
        except Exception as e:
            st.error(f"Error reading image: {e}")

    col1, col2 = st.columns(2)
    with col1:
        sat_lat = st.number_input("Latitude", value=26.912400, format="%.6f", key="insights_lat")
    with col2:
        sat_lon = st.number_input("Longitude", value=75.787300, format="%.6f", key="insights_lon")

    if st.button("🔮 Generate Geographic Insights", use_container_width=True, type="primary"):
        with st.spinner("📡 Fetching multi-source satellite imagery…"):
            try:
                images_dict = get_all_satellite_imagery(sat_lat, sat_lon)
                st.session_state["insights_images"] = images_dict
            except Exception as exc:
                st.error(f"Failed to fetch imagery: {exc}")
                images_dict = None

        if images_dict:
            # Display all fetched satellite images
            st.markdown("#### Satellite Feeds")
            cols = st.columns(len(images_dict))
            for i, (provider, img) in enumerate(images_dict.items()):
                with cols[i]:
                    if isinstance(img, Image.Image):
                        st.image(img, caption=provider, use_container_width=True)
                    else:
                        st.error(f"{provider}\n{img}", icon="❌")

            with st.spinner("🤖 Generative AI analyzing geography & surroundings…"):
                try:
                    # If we have a ground image, add it to the image dict passed to Gemini
                    if ground_image:
                        images_dict["Ground-Level Property Photo"] = ground_image
                        
                    report = run_multi_source_analysis(GEMINI_API_KEY, images_dict, deed_data=None)
                    st.success("Analysis Complete!")
                    st.markdown("### 📋 AI Insights Report")
                    st.markdown(f"> {report.replace(chr(10), chr(10) + '> ')}")
                except Exception as exc:
                    st.error(f"AI analysis failed: {exc}")

    st.markdown('</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# MODE 2: BATCH PROCESSING
# ═══════════════════════════════════════════════════════════════════════════════
elif mode == "📦 Batch Processing":

    st.markdown('<div class="avm-card">', unsafe_allow_html=True)
    st.markdown('<span class="badge badge-blue">Step 1</span>', unsafe_allow_html=True)
    st.markdown("### 📦 Upload Multiple Property Documents")
    st.markdown(
        "Upload **multiple scanned property deed PDFs** — each will be processed "
        "individually with confidence scoring, and all reports packaged into a single zip."
    )

    batch_pdfs = st.file_uploader(
        "Drop all PDFs here",
        type=["pdf"],
        accept_multiple_files=True,
        label_visibility="collapsed",
        key="batch_upload",
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if batch_pdfs and len(batch_pdfs) > 0:


        st.markdown('<span class="badge badge-amber">Step 2</span>', unsafe_allow_html=True)
        st.markdown(f"### 🤖 Batch Extraction — {len(batch_pdfs)} Document{'s' if len(batch_pdfs) > 1 else ''}")

        # Bank selector for batch
        bank_names = list(BANK_CONFIGS.keys())
        batch_bank = st.selectbox(
            "Select Bank Template for all reports",
            bank_names,
            key="batch_bank_selector",
        )

        if not GEMINI_API_KEY:
            st.error("AI Engine not configured. Please contact the administrator.", icon="🚨")
        else:
            if st.button("⚡ Process Batch", use_container_width=True, key="batch_extract"):
                all_results = []
                progress_bar = st.progress(0, text="Starting batch extraction…")

                for idx, pdf_file in enumerate(batch_pdfs):
                    file_label = pdf_file.name
                    progress_bar.progress(
                        idx / len(batch_pdfs),
                        text=f"📄 Processing {file_label} ({idx + 1}/{len(batch_pdfs)})…"
                    )

                    try:
                        pdf_bytes = pdf_file.read()
                        page_images = pdf_to_pil_images(pdf_bytes, dpi=300)
                        extracted_data, raw_response = extract_with_gemini(
                            GEMINI_API_KEY, page_images, DEFAULT_MODEL
                        )
                        extracted_data["_source_file"] = file_label
                        extracted_data["_page_count"] = len(page_images)
                        all_results.append(extracted_data)
                    except Exception as exc:
                        st.warning(f"⚠️ Failed to extract from {file_label}: {exc}")
                        all_results.append({
                            "_source_file": file_label,
                            "_error": str(exc),
                        })

                progress_bar.progress(1.0, text="✅ All documents processed!")

                if all_results:
                    st.session_state["batch_results"] = all_results
                    st.session_state["batch_bank"] = batch_bank
                    successful = sum(1 for r in all_results if "_error" not in r)
                    st.success(f"✅ Extracted data from {successful}/{len(all_results)} documents!", icon="📦")

        st.markdown('</div>', unsafe_allow_html=True)

        # ── Batch Results ─────────────────────────────────────────────────────
        if st.session_state.get("batch_results"):
            batch_results = st.session_state["batch_results"]
            batch_bank = st.session_state.get("batch_bank", list(BANK_CONFIGS.keys())[0])

        
            st.markdown('<span class="badge badge-green">Results</span>', unsafe_allow_html=True)
            st.markdown("## 📊 Batch Extraction Results")

            import pandas as pd

            # Show each result in an expander
            for i, result in enumerate(batch_results):
                source = result.get("_source_file", f"Document {i+1}")

                if "_error" in result:
                    with st.expander(f"❌ {source} — FAILED", expanded=False):
                        st.error(f"Extraction failed: {result['_error']}")
                    continue

                customer = get_val(result, "customer_name")
                cust_conf = get_conf(result, "customer_name")
                with st.expander(f"{conf_badge(cust_conf).split()[0]} {source} — {customer}", expanded=(i == 0)):
            
                    st.markdown("#### 👤 Customer & Property Summary")
                    st.table(pd.DataFrame([
                        {"Field": "Customer Name", "English": get_val(result, "customer_name"), "Hindi": get_hindi(result, "customer_name"), "Conf.": conf_badge(get_conf(result, "customer_name"))},
                        {"Field": "Full Address", "English": get_val(result, "address"), "Hindi": get_hindi(result, "address"), "Conf.": conf_badge(get_conf(result, "address"))},
                        {"Field": "Land Area", "English": get_val(result, "land_area"), "Hindi": get_hindi(result, "land_area"), "Conf.": conf_badge(get_conf(result, "land_area"))},
                    ]))
                    st.markdown('</div>', unsafe_allow_html=True)

            
                    st.markdown("#### 🧭 Boundaries & Dimensions")
                    st.table(pd.DataFrame([
                        {"Dir.": "East", "Dim. (EN)": get_val(result, "dim_east"), "Dim. (HI)": get_hindi(result, "dim_east"), "Boundary (EN)": get_val(result, "bound_east"), "Boundary (HI)": get_hindi(result, "bound_east"), "Conf.": conf_badge(get_conf(result, "dim_east"))},
                        {"Dir.": "West", "Dim. (EN)": get_val(result, "dim_west"), "Dim. (HI)": get_hindi(result, "dim_west"), "Boundary (EN)": get_val(result, "bound_west"), "Boundary (HI)": get_hindi(result, "bound_west"), "Conf.": conf_badge(get_conf(result, "dim_west"))},
                        {"Dir.": "North", "Dim. (EN)": get_val(result, "dim_north"), "Dim. (HI)": get_hindi(result, "dim_north"), "Boundary (EN)": get_val(result, "bound_north"), "Boundary (HI)": get_hindi(result, "bound_north"), "Conf.": conf_badge(get_conf(result, "dim_north"))},
                        {"Dir.": "South", "Dim. (EN)": get_val(result, "dim_south"), "Dim. (HI)": get_hindi(result, "dim_south"), "Boundary (EN)": get_val(result, "bound_south"), "Boundary (HI)": get_hindi(result, "bound_south"), "Conf.": conf_badge(get_conf(result, "dim_south"))},
                    ]))
                    st.markdown('</div>', unsafe_allow_html=True)

            # ── Zip Download ──────────────────────────────────────────────────
            st.markdown("### 📥 Download All Reports")
            successful_results = [r for r in batch_results if "_error" not in r]

            if successful_results:
                try:
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                        for result in successful_results:
                            customer = get_val(result, "customer_name")
                            fname = f"{sanitize_filename(customer)}_Valuation.xlsx"
                            excel_data = generate_bank_report(batch_bank, result)
                            zf.writestr(fname, excel_data.getvalue())

                    zip_buffer.seek(0)
                    st.download_button(
                        label=f"📦 Download All {len(successful_results)} Reports (.zip)",
                        data=zip_buffer,
                        file_name="batch_valuation_reports.zip",
                        mime="application/zip",
                        type="primary",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Failed to generate zip: {e}")
            else:
                st.warning("No successful extractions to download.")

    else:
        # Idle hero state for Batch mode

        st.markdown(
            """
            <div style="text-align:center; padding: 3rem 1rem;">
                <div style="font-size:4rem; margin-bottom:1rem;">📦</div>
                <h2 style="color:#7eb8f7; margin-bottom:0.5rem;">Batch Processing</h2>
                <p style="color:#607d8b; max-width:520px; margin:0 auto;">
                    Upload multiple property deeds at once. The AI will extract data
                    from each document with confidence scoring, and package all
                    bank valuation reports into a single downloadable zip file.
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown('</div>', unsafe_allow_html=True)

        feat1, feat2, feat3 = st.columns(3)
        for col, icon, title, desc in [
            (feat1, "📄", "Multi-Upload", "Process 2-50 deeds in one click"),
            (feat2, "🎯", "Confidence Scores", "🟢🟡🔴 badges show AI certainty"),
            (feat3, "📦", "Zip Download", "All reports packaged into one file"),
        ]:
            with col:
                st.markdown(
                    f"""
                    <div class="avm-card" style="text-align:center; padding:1.2rem;">
                        <div style="font-size:2rem;">{icon}</div>
                        <div style="font-weight:600; color:#7eb8f7; margin:0.4rem 0 0.2rem;">{title}</div>
                        <div style="font-size:0.82rem; color:#7a90a8;">{desc}</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )


# ═══════════════════════════════════════════════════════════════════════════════
# MODE 3: CHAIN OF TITLE
# ═══════════════════════════════════════════════════════════════════════════════
elif mode == "🔗 Chain of Title":

    st.markdown('<div class="avm-card">', unsafe_allow_html=True)
    st.markdown('<span class="badge badge-blue">Step 1</span>', unsafe_allow_html=True)
    st.markdown("### 🔗 Upload All Deeds in the Chain")
    st.markdown(
        "Upload **all property documents** in the ownership chain — sale deeds, "
        "allotment letters, gift deeds, partition deeds, etc. The AI will extract "
        "data from each and build the ownership timeline."
    )

    uploaded_pdfs = st.file_uploader(
        "Drop all PDFs here",
        type=["pdf"],
        accept_multiple_files=True,
        label_visibility="collapsed",
        key="chain_upload",
    )
    st.markdown("</div>", unsafe_allow_html=True)

    if uploaded_pdfs and len(uploaded_pdfs) > 0:


        st.markdown('<span class="badge badge-amber">Step 2</span>', unsafe_allow_html=True)
        st.markdown(f"### 🤖 AI Extraction — {len(uploaded_pdfs)} Document{'s' if len(uploaded_pdfs) > 1 else ''}")

        if not GEMINI_API_KEY:
            st.error("AI Engine not configured. Please contact the administrator.", icon="🚨")
        else:
            if st.button("⚡ Extract All & Build Chain", use_container_width=True, key="chain_extract"):
                all_extractions = []
                progress_bar = st.progress(0, text="Starting extraction…")

                for idx, pdf_file in enumerate(uploaded_pdfs):
                    file_label = pdf_file.name
                    progress_bar.progress(
                        (idx) / len(uploaded_pdfs),
                        text=f"📄 Processing {file_label} ({idx + 1}/{len(uploaded_pdfs)})…"
                    )

                    try:
                        pdf_bytes = pdf_file.read()
                        page_images = pdf_to_pil_images(pdf_bytes, dpi=300)
                        extracted_data, _ = extract_with_gemini(
                            GEMINI_API_KEY, page_images, DEFAULT_MODEL
                        )
                        extracted_data["_source_file"] = file_label
                        extracted_data["_page_count"] = len(page_images)
                        all_extractions.append(extracted_data)
                    except Exception as exc:
                        st.warning(f"⚠️ Failed to extract from {file_label}: {exc}")

                progress_bar.progress(1.0, text="✅ All documents processed!")

                if all_extractions:
                    chain = build_ownership_chain(all_extractions)
                    st.session_state["chain_data"] = chain
                    st.session_state["chain_extractions"] = all_extractions
                    st.success(f"✅ Extracted data from {len(all_extractions)} documents and built ownership chain!", icon="🔗")
                else:
                    st.error("No documents were successfully extracted.")

        st.markdown("</div>", unsafe_allow_html=True)

        # ── Chain of Title Results ────────────────────────────────────────────
        if st.session_state.get("chain_data"):
            chain = st.session_state["chain_data"]
            all_extractions = st.session_state.get("chain_extractions", [])

        
            st.markdown('<span class="badge badge-green">Results</span>', unsafe_allow_html=True)
            st.markdown("## 🔗 Chain of Title — Ownership Timeline")

            tab_timeline, tab_details, tab_json = st.tabs(
                ["📜 Ownership Timeline", "📋 Individual Deed Details", "🗂️ All Data (JSON)"]
            )

            with tab_timeline:
                st.markdown('<div class="avm-card">', unsafe_allow_html=True)
                st.markdown(f"**{len(chain)} documents** in the ownership chain")

                # Summary of chain
                if chain:
                    first_buyer = get_val(chain[0]["data"], "customer_name")
                    last_buyer = get_val(chain[-1]["data"], "customer_name")
                    first_date = chain[0]["date_str"]
                    last_date = chain[-1]["date_str"]
                    st.markdown(
                        f"**Original owner:** {first_buyer} ({first_date})  \n"
                        f"**Current owner:** {last_buyer} ({last_date})"
                    )

                timeline_html = render_chain_timeline(chain)
                st.markdown(timeline_html, unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)

            with tab_details:
                for i, entry in enumerate(chain):
                    deed = entry["data"]
                    source = deed.get("_source_file", f"Document {i+1}")
                    badge = " ★ CURRENT" if entry["is_current"] else ""
                    st.markdown(f'<div class="avm-card{" current" if entry["is_current"] else ""}">', unsafe_allow_html=True)
                    st.markdown(f"#### 📄 {source}{badge}")
                    st.markdown(f"**Date:** {entry['date_str']}")

                    # Render all categories for this deed
                    st.markdown("#### 👤 Customer & Property Summary")
                    import pandas as pd
                    st.table(pd.DataFrame([
                        {"Field": "Customer Name", "English": get_val(deed, "customer_name"), "Hindi": get_hindi(deed, "customer_name"), "Conf.": conf_badge(get_conf(deed, "customer_name"))},
                        {"Field": "Full Address", "English": get_val(deed, "address"), "Hindi": get_hindi(deed, "address"), "Conf.": conf_badge(get_conf(deed, "address"))},
                        {"Field": "Land Area", "English": get_val(deed, "land_area"), "Hindi": get_hindi(deed, "land_area"), "Conf.": conf_badge(get_conf(deed, "land_area"))},
                    ]))

                    st.markdown("#### 🧭 Boundaries & Dimensions")
                    st.table(pd.DataFrame([
                        {"Dir.": "East", "Dim. (EN)": get_val(deed, "dim_east"), "Dim. (HI)": get_hindi(deed, "dim_east"), "Boundary (EN)": get_val(deed, "bound_east"), "Boundary (HI)": get_hindi(deed, "bound_east"), "Conf.": conf_badge(get_conf(deed, "dim_east"))},
                        {"Dir.": "West", "Dim. (EN)": get_val(deed, "dim_west"), "Dim. (HI)": get_hindi(deed, "dim_west"), "Boundary (EN)": get_val(deed, "bound_west"), "Boundary (HI)": get_hindi(deed, "bound_west"), "Conf.": conf_badge(get_conf(deed, "dim_west"))},
                        {"Dir.": "North", "Dim. (EN)": get_val(deed, "dim_north"), "Dim. (HI)": get_hindi(deed, "dim_north"), "Boundary (EN)": get_val(deed, "bound_north"), "Boundary (HI)": get_hindi(deed, "bound_north"), "Conf.": conf_badge(get_conf(deed, "dim_north"))},
                        {"Dir.": "South", "Dim. (EN)": get_val(deed, "dim_south"), "Dim. (HI)": get_hindi(deed, "dim_south"), "Boundary (EN)": get_val(deed, "bound_south"), "Boundary (HI)": get_hindi(deed, "bound_south"), "Conf.": conf_badge(get_conf(deed, "dim_south"))},
                    ]))

                    st.markdown('</div>', unsafe_allow_html=True)
                    st.markdown("")

            with tab_json:
                st.markdown("#### All Extracted Data")
                st.json(all_extractions)

    else:
        # Idle hero state for Chain mode
        st.markdown('<div class="avm-card">', unsafe_allow_html=True)
        st.markdown(
            """
            <div style="text-align:center; padding: 3rem 1rem;">
                <div style="font-size:4rem; margin-bottom:1rem;">🔗</div>
                <h2 style="color:#7eb8f7; margin-bottom:0.5rem;">Chain of Title</h2>
                <p style="color:#607d8b; max-width:520px; margin:0 auto;">
                    Upload all property deeds in the ownership chain above.
                    The AI will extract data from each document, match buyer→seller
                    names to build the ownership timeline, detect any gaps,
                    and highlight the current deed.
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown("</div>", unsafe_allow_html=True)

        feat1, feat2, feat3 = st.columns(3)
        for col, icon, title, desc in [
            (feat1, "📚", "Multi-Document",  "Upload 2-20 deeds at once"),
            (feat2, "🔗", "Auto Chain Link", "Matches buyer→seller across deeds"),
            (feat3, "⚠️", "Gap Detection",   "Flags breaks in the ownership chain"),
        ]:
            with col:
                st.markdown(
                    f"""
                    <div class="avm-card" style="text-align:center; padding:1.2rem;">
                        <div style="font-size:2rem;">{icon}</div>
                        <div style="font-weight:600; color:#7eb8f7; margin:0.4rem 0 0.2rem;">{title}</div>
                        <div style="font-size:0.82rem; color:#7a90a8;">{desc}</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

# ═══════════════════════════════════════════════════════════════════════════════
# MODE 5: FORMAT CONVERTER
# ═══════════════════════════════════════════════════════════════════════════════
elif mode == "🔄 Format Converter":
    st.markdown('<div class="avm-card">', unsafe_allow_html=True)
    st.markdown("## 🔄 Format Shifting Engine")
    st.markdown("Upload a scanned PDF valuation report. AI will extract 50+ structural data points and safely inject them into the beautiful **Axis Bank Excel template**.")
    
    uploaded_pdf = st.file_uploader("Upload Scanned Report (.pdf)", type=["pdf"], key="converter_upload")
    
    if uploaded_pdf:
        if st.button("🚀 Run Format Conversion", use_container_width=True):
            # 1. Convert PDF to Images
            with st.spinner("📄 Rasterizing PDF into high-res images..."):
                pdf_bytes = uploaded_pdf.read()
                # Limit to 5 pages natively like the converter does
                page_images = pdf_to_pil_images(pdf_bytes, dpi=150)[:5]
            
            # 2. Extract Data via Gemini
            extracted_data, raw_text = extract_format_converter_with_gemini(GEMINI_API_KEY, page_images)
            
            if extracted_data:
                st.success("✅ Gemini Data Extraction Complete!")
                with st.expander("🔍 View Extracted Data Mapping", expanded=False):
                    st.json(extracted_data)
                
                # 3. Inject into Excel
                with st.spinner("📊 Mapping data to Axis Template cells..."):
                    excel_buffer = generate_axis_report(extracted_data)
                
                # 4. Success Dashboard & Download Button
                st.markdown("---")
                st.markdown("### 🎯 Conversion Successful")
                st.download_button(
                    label="📥 Download Injected Axis Template (.xlsx)",
                    data=excel_buffer,
                    file_name=f"{sanitize_filename(extracted_data.get('owner_name', 'Converted_Report'))}_Axis_Format.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
    st.markdown('</div>', unsafe_allow_html=True)
