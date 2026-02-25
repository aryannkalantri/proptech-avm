"""
create_chola_template.py
------------------------
Recreates the Cholamandalam Excel valuation template.
Run once: python create_chola_template.py
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def create_chola_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valuation Report"

    # ── Styles ────────────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=14, color="1B4F72")
    section_font = Font(name="Calibri", bold=True, size=11, color="1B4F72")
    label_font = Font(name="Calibri", bold=True, size=10)
    value_font = Font(name="Calibri", size=10)
    header_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    section_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 25

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    ws["A1"] = "CHOLAMANDALAM — PROPERTY VALUATION REPORT"
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = header_fill

    ws.merge_cells("A2:E2")
    ws["A2"] = "Prepared by Antigravity PropTech Vision Extractor"
    ws["A2"].font = Font(name="Calibri", italic=True, size=9, color="7F8C8D")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Section: Customer Details ─────────────────────────────────────────────
    ws.merge_cells("A4:E4")
    ws["A4"] = "1. CUSTOMER DETAILS"
    ws["A4"].font = section_font
    ws["A4"].fill = section_fill

    ws["A5"] = "Customer Name"
    ws["A5"].font = label_font
    ws["B5"] = ""  # ← filled by app
    ws["B5"].font = value_font

    # ── Section: Property Location ────────────────────────────────────────────
    ws.merge_cells("A7:E7")
    ws["A7"] = "2. PROPERTY LOCATION"
    ws["A7"].font = section_font
    ws["A7"].fill = section_fill

    # Row 8-9: labels
    ws["A8"] = "City/Town/Village"
    ws["A8"].font = label_font
    ws["A9"] = "District"
    ws["A9"].font = label_font
    ws["A10"] = "State"
    ws["A10"].font = label_font

    # ── Section: Property Details ─────────────────────────────────────────────
    ws.merge_cells("A12:E12")
    ws["A12"] = "3. PROPERTY DETAILS"
    ws["A12"].font = section_font
    ws["A12"].fill = section_fill

    ws["A13"] = "Type of Property"
    ws["A13"].font = label_font
    ws["A14"] = "Address (as per document)"
    ws["A14"].font = label_font
    ws["B14"] = ""  # ← filled by app
    ws["A15"] = "Address (as per site)"
    ws["A15"].font = label_font
    ws["B15"] = ""  # ← filled by app

    # ── Section: Land Details ─────────────────────────────────────────────────
    ws.merge_cells("A17:E17")
    ws["A17"] = "4. LAND DETAILS"
    ws["A17"].font = section_font
    ws["A17"].fill = section_fill

    ws["A18"] = "Survey No. / Plot No."
    ws["A18"].font = label_font
    ws["A19"] = "T.S. No."
    ws["A19"].font = label_font
    ws["A20"] = "Door No."
    ws["A20"].font = label_font

    ws["A22"] = "Land Area (as per document)"
    ws["A22"].font = label_font
    ws["B22"] = ""  # ← filled by app
    ws["C22"] = ""  # ← filled by app

    # ── Section: Dimensions ───────────────────────────────────────────────────
    ws.merge_cells("A25:E25")
    ws["A25"] = "5. DIMENSIONS & BOUNDARIES"
    ws["A25"].font = section_font
    ws["A25"].fill = section_fill

    # Dimension headers
    for col, direction in [("B", "East"), ("C", "West"), ("D", "North"), ("E", "South")]:
        ws[f"{col}32"] = direction
        ws[f"{col}32"].font = label_font
        ws[f"{col}32"].alignment = Alignment(horizontal="center")
        ws[f"{col}32"].fill = header_fill
        ws[f"{col}32"].border = thin_border

    ws["A32"] = "Direction"
    ws["A32"].font = label_font
    ws["A32"].fill = header_fill
    ws["A32"].border = thin_border

    # Dimension rows
    ws["A33"] = "Dimension (as per document)"
    ws["A33"].font = label_font
    ws["A33"].border = thin_border
    ws["A34"] = "Dimension (as per site)"
    ws["A34"].font = label_font
    ws["A34"].border = thin_border

    for col in ["B", "C", "D", "E"]:
        for row in [33, 34]:
            ws[f"{col}{row}"] = ""  # ← filled by app
            ws[f"{col}{row}"].font = value_font
            ws[f"{col}{row}"].border = thin_border
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center")

    # Boundary rows
    ws["A35"] = ""
    ws["A36"] = "Boundary (as per document)"
    ws["A36"].font = label_font
    ws["A36"].border = thin_border
    ws["A37"] = "Boundary (as per site)"
    ws["A37"].font = label_font
    ws["A37"].border = thin_border

    for col in ["B", "C", "D", "E"]:
        for row in [36, 37]:
            ws[f"{col}{row}"] = ""  # ← filled by app
            ws[f"{col}{row}"].font = value_font
            ws[f"{col}{row}"].border = thin_border
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center")

    # ── Save ──────────────────────────────────────────────────────────────────
    os.makedirs("templates", exist_ok=True)
    wb.save("templates/chola_template.xlsx")
    print("✅ Chola template saved to templates/chola_template.xlsx")


if __name__ == "__main__":
    create_chola_template()
